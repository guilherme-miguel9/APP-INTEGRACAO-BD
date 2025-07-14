import pandas as pd
import psycopg2
from psycopg2 import extras
from psycopg2 import sql
import ttkbootstrap as ttk
import tkinter as tk
from tkinter import filedialog
import os
import configparser
from sqlalchemy import create_engine, inspect, extract
from urllib.parse import quote_plus
import threading
import sys
import re
from openpyxl import load_workbook
import time

#Configurações e váriaveis globais

#Configurações do banco de dados
ini_path = os.path.join(os.path.dirname(__file__), 'config.ini')
config = configparser.ConfigParser()
config.read(ini_path)

db_host = config['database']['host']
db_user = config['database']['user']
db_password = config['database']['password']
db_name = config['database']['dbname']

db_password_escaped = quote_plus(db_password)
engine = create_engine(f'postgresql://{db_user}:{db_password_escaped}@{db_host}:5432/{db_name}')
insp = inspect(engine)

#Configurações do tKinter
root = ttk.Window(themename="pulse")
root.title("SCRIPT CONEXÃO AO BANCO DE DADOS")
root.geometry("1300x800")

dataframe = []
parar_barra_progresso = threading.Event()
tempo_medio_por_linha = 0.05
ponteiro_01 = False

#Funções abaixo 

def selecionar_arquivo_excel():
    global caminho_arquivo
    global nome_arquivo
    global abas

    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione um arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    nome_arquivo = os.path.basename(caminho_arquivo)
    
    if lista_box.size() > 0:
        lista_box.delete(0, tk.END)  
        lista_box.insert(tk.END, nome_arquivo)
    else:
        lista_box.insert(tk.END, nome_arquivo)

    #Variável para abrir os nomes de sheets do Excel
    abas = pd.ExcelFile(caminho_arquivo)   #EXEMPLO: ['CR12','CR13','CR14E15']

    return caminho_arquivo

def threading_processar_excel():
    botao_abrir_excel['state'] = 'disabled'
    botao_escolher_excel['state'] = 'disabled'
    parar_barra_progresso.clear()
    #contar_linhas = contar_linhas_excel()
    threading.Thread(target=processar_excel, daemon=True).start()
    threading.Thread(target=atualizar_porcentagem_progresso(), daemon=True).start()
    
    

def atualizar_box_tabelas():
    todas_tabelas = insp.get_table_names(schema=schema_box.get())
    tabela_box.config(values=todas_tabelas)
    tabela_box.set('')

def atualizar_porcentagem_progresso():    
    tempo_por_linha = tempo_medio_por_linha * 100000
    inicio = time.time()
    progress['value'] = 0

    while not parar_barra_progresso.is_set():
        tempo_decorrido = time.time() - inicio
        progresso_barra = min(int((tempo_decorrido / tempo_por_linha) * 100), 99)
        progress['value'] = progresso_barra
        porcentagem_progresso = int(progress["value"])
        label_porcentagem.config(text=f"{porcentagem_progresso}%")
        #root.update_idletasks()
        time.sleep(0.1)
        progress.update_idletasks()
        label_porcentagem.update_idletasks()
        if ponteiro_01 == True:
            break

    progress['value'] = 100
    label_porcentagem.config(text="100%")
    root.update_idletasks()
    
# def contar_linhas_excel():
#     wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
#     return (sum(wb[sheet].max_row for sheet in wb.sheetnames))
    
    
#########################################################

colunasarrumadas = {
        'Nº' : str,
        'Nº item da ordem' : str, 
        'Instal' : str, 
        'Registrador' : str, 
        'Nº da casa': str,
        'Sequência' : str,
        'Contrato' : str,
        'ObjLigacao' : str,
        'Nº Poste' : str,
        'Nº Serie' : str,
        'Unid.leit' : str,
        'Cta.contr.' : str,
        'Coment.leitura' : str
}

mapeamento_colunas = {
    'Nome da Origem.1' : 'Data_Atual',
    'Nº' : 'N',
    'Nº item da ordem' : 'Numero_Item_Ordem',
    'Instal' : 'Instalacao',
    'Registrador' : 'Registrador',
    'Rua': 'Rua',
    'Nº da casa' : 'N_casa',
    'Sequência' : 'Sequencia',
    'Contrato' : 'Contrato',
    'Latitude localiz.geográfica' : 'Latitude',
    'Longitude localiz.geográfica' : 'Longitude',
    'Val Fat' : 'Valor_fatura',
    'NomeCliente' : 'Nome_Cliente',
    'Complemento' : 'Complemento',
    'Ponto Ref' : 'Ponto_Ref',
    'Local' : 'Municipio',
    'Bairro' : 'Bairro',
    'Sigla edifício' : 'Sigla_Edificio',
    'Nº sala' : 'N_sala',
    'Andar' : 'Andar',
    'Complemento endereco' : 'Complemento_Endereco',
    'ObjLigacao' : 'Objeto_Ligacao',
    'Nº Poste' : 'N_poste',
    'Nº Serie' : 'N_serie',
    'Unid.leit' : 'Unidade_Leitura',
    'O. leitura real' : 'O_Leitura_Real',
    'O. Sem leit real' : 'O_Sem_Leitura_Real',
    'Nota leit.' : 'Nota_Leitura',
    'Hora leit.' : 'Hora_Leitura',
    'Seq.Mod' : 'SeqMod',
    'Cond WOL' : 'CondWOL',
    'Leit' : 'Codigo_Leitor',
    'Nome leit' : 'Nome_Leit',
    'Indic Foto' : 'Indicador_Foto',
    'Interv.Leit' : 'Intervalo_leitura',
    'Cta.contr.' : 'Conta_Contrato',
    'Abaixo lim' : 'Abaixo_Lim',
    'Excede lim' : 'Excede_Lim',
    'Desvio leit' : 'Desvio_Leitura',
    'Fat. Assin' : 'Fat_Assin',
    'Coment.leitura' : 'Comentario_Leitura',
    'Coment.fatura' : 'Comentario_Fatura',
    'Tipo rota' : 'Tipo_Rota',
    'Tipo ordem' : 'Tipo_Ordem',
    'Impresso' : 'Impresso',
    'ResCampo' : 'Res_campo',
    'FA CT OK' : 'FACT_OK'
}

#########################################################
def processar_excel():
    

    

    ano_tabela_box = re.search(r'\d+', tabela_box.get())

    dicionario_meses = {
        "jan_" + ano_tabela_box.group() : 1,
        "fev_" + ano_tabela_box.group() : 2,
        "mar_" + ano_tabela_box.group() : 3,
        "abr_" + ano_tabela_box.group() : 4,
        "mai_" + ano_tabela_box.group() : 5,
        "jun_" + ano_tabela_box.group() : 6,
        "jul_" + ano_tabela_box.group() : 7,
        "ago_" + ano_tabela_box.group() : 8,
        "set_" + ano_tabela_box.group() : 9,
        "out_" + ano_tabela_box.group() : 10,
        "nov_" + ano_tabela_box.group() : 11,
        "dez_" + ano_tabela_box.group() : 12,
        "teste_" + ano_tabela_box.group() : 5
    }

    root.update_idletasks()
    progress.update_idletasks()
    label_porcentagem.update_idletasks()

    try:
        #progress.start()
        

        mensagem.config(text=f"Processando o arquivo Excel... {nome_arquivo}", foreground="blue")

        
        tabela_leitura = tabela_box.get()
        schema_leitura = schema_box.get()

        #Concatenar os dados de todas as abas em um unico df
        df = pd.concat([pd.read_excel(caminho_arquivo, sheet_name=aba, engine='openpyxl', dtype=colunasarrumadas) for aba in abas.sheet_names], ignore_index=True)

        ponteiro_01 = True
        
        df = df.fillna('')

        df.rename(columns=mapeamento_colunas, inplace=True)

        df['Hora_Leitura'] = pd.to_datetime(df['Hora_Leitura'], format='%H:%M:%S').dt.time

        df['Intervalo_leitura'] = pd.to_datetime(df['Intervalo_leitura'], format='%H:%M:%S').dt.time

        df['Data_Atual'] = pd.to_datetime(df['Data_Atual'], format='%d.%m.%Y', errors='coerce')

        df['Latitude'] = df['Latitude'].astype(float)

        df['Longitude'] = df['Longitude'].astype(float)

        df['Valor_fatura'] = df['Valor_fatura'].astype(float)

        dias = df['Data_Atual'].dt.day.dropna().unique().tolist()

        if tabela_box.get() in dicionario_meses:
            mes_tabela_box = dicionario_meses[tabela_box.get()]

        mes_excel = df['Data_Atual'].iloc[0].month

        if mes_excel != mes_tabela_box:
            sys.exit(f"O mês da tabela selecionada {tabela_box.get()} não corresponde ao mês do Excel selecionado {nome_arquivo}. Por favor, selecione a tabela correta.")
        
        mensagem.config(text=f"Arquivo do Excel processo com sucesso!", foreground="green")
        

    except SystemExit as e:
        mensagem.config(text=f"ERROR: {e}", foreground="red")
        parar_barra_progresso.set()
        botao_abrir_excel['state'] = 'normal'
        botao_escolher_excel['state'] = 'normal'
        return

    except Exception as e:
        mensagem.config(text=f"Ocorreu um erro ao abrir o Excel: {e}", foreground="red")
        parar_barra_progresso.set()
        botao_abrir_excel['state'] = 'normal'
        botao_escolher_excel['state'] = 'normal'
        return

    try:
        
        mensagem.config(text="Conectando ao banco de dados...", foreground="blue")

        with psycopg2.connect(
            dbname=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port="5432"
        ) as conn:
            conn.set_client_encoding('UTF8')

            
            with conn.cursor() as cursor:
                #Setar o estilo da data
                cursor.execute("SET datestyle TO 'ISO, DMY';")
                
                query_dias = sql.SQL("DELETE FROM {schema_leitura}.{tabela_leitura} WHERE EXTRACT(DAY FROM data_atual) IN ({placeholders})").format(
                    tabela_leitura=sql.Identifier(tabela_leitura),
                    schema_leitura=sql.Identifier(schema_leitura),
                    placeholders=sql.SQL(',').join(sql.Placeholder() * len(dias))
                    )
                
                cursor.execute(query_dias, (dias))

                #Inserir o dataframe em uma lista de tuplas
                data_to_insert = [tuple(row) for row in df.itertuples(index=False, name=None)]
                conn.commit()
                #modificar a tabela que vai ser inserida no BANCO DE DADOS e inserir os dados
                query = f"INSERT INTO {schema_leitura}.{tabela_leitura} ({', '.join(df.columns)}) VALUES ({', '.join(['%s']*len(df.columns))})"


                mensagem_quantidade.config(text=f"Quantidade de registros: {len(data_to_insert)}", foreground="purple")
            

                extras.execute_batch(cursor, query, data_to_insert, page_size=50000)

                conn.commit()
                mensagem.config(text="Banco de dados atualizado com sucesso!", foreground="green")
        
    except Exception as e:
        mensagem.config(text=f"Ocorreu um erro ao entrar no banco de dados: {e}", foreground="red")
        parar_barra_progresso.set()
        botao_abrir_excel['state'] = 'normal'
        botao_escolher_excel['state'] = 'normal'
        return

    finally:
        if cursor:
            cursor.close()

        if conn:
            conn.close()
            parar_barra_progresso.set()
            botao_abrir_excel['state'] = 'normal'
            botao_escolher_excel['state'] = 'normal'
#########################################################

# Criação da interface gráfica

#BLOCO DO BANCO DE DADOS
bloco_bd = ttk.LabelFrame(root, text="Configurações do Banco de Dados", bootstyle="info")
bloco_bd.pack(side="top", anchor="n", padx=20, pady=20, expand=True)

botao_escolher_excel = ttk.Button(bloco_bd, text="Clique para escolher o Excel: ", command=selecionar_arquivo_excel)
botao_escolher_excel.pack(pady=10)

lista_box = tk.Listbox(bloco_bd, width=30, height=1, font=("Arial", 10))
lista_box.pack(pady=2.5)

mensagem_schema = ttk.Label(bloco_bd, text="Escolha a pasta e logo após a tabela para inserir os dados do Excel", bootstyle="secondary")
mensagem_schema.pack(pady=10)
todas_schema = insp.get_schema_names()
schema_box = ttk.Combobox(bloco_bd, values=todas_schema, bootstyle="secondary", width=20, justify="center")
schema_box.current(0)
schema_box.pack(pady=10)

todas_tabelas = insp.get_table_names(schema=schema_box.get())
tabela_box = ttk.Combobox(bloco_bd, values=[], bootstyle="secondary", heigh=5, width=20, justify="center") 

tabela_box.pack(pady=10)

schema_box.bind("<<ComboboxSelected>>", lambda e: atualizar_box_tabelas())

########################
#AREA DE MENSAGENS
bloco_mensagem = ttk.LabelFrame(root, text="Mensagens", bootstyle="info")
bloco_mensagem.pack(anchor="center", padx=20, pady=20)

mensagem = ttk.Label(bloco_mensagem, text="", wraplength=400, anchor="w", justify="left")
mensagem.pack(padx=20, pady=20)

mensagem_quantidade = ttk.Label(bloco_mensagem, text="Quantidade de registros: ", bootstyle="secondary")
mensagem_quantidade.pack(pady=5)

#BOTAO PARA ABRIR O EXCEL
botao_abrir_excel = ttk.Button(bloco_bd, text="Abrir Excel", command=threading_processar_excel, bootstyle="primary")
botao_abrir_excel.pack(pady=10)

#BARRA DE PROGRESSO
label_porcentagem = ttk.Label(bloco_bd, text="0%", bootstyle="dark")
label_porcentagem.pack(pady=5)

progress = ttk.Progressbar(bloco_bd, orient="horizontal", length=200, mode="determinate", bootstyle="success")
progress.pack(pady=10)

root.mainloop()